# Importando a biclioteca para minupulação de excel\n",
import openpyxl
import shutil
   
    # Buscando os módulos para leitura e criação de planilhas
from openpyxl import workbook, load_workbook
from pathlib import Path
    
    # Criando as variáveis que receberão o local dos arquivos de dados e modelo de simulação\n
local_arquivo_dados = 'C:/Users/pedro/Desktop/Automatizacao/Dados.xlsx' # os inputs entram aqui
local_arquivo_simulacao = 'C:/Users/pedro/Desktop/Automatizacao/Simulacao_Prebeg.xlsx' # os inputs entram aqui
local_arquivo_salarios = 'C:/Users/pedro/Desktop/Automatizacao/Salarios.xlsx' # os inputs entram aqui
    
    # Criando as variáveis que receberão os dados dos arquivos completos
wb = load_workbook(local_arquivo_dados)
wb1 = load_workbook(local_arquivo_simulacao)
wb2 = load_workbook(local_arquivo_salarios)
    #dados = wb.get_sheet_by_name('DadosParticipante')
    
    # Criando as variáveis que receberão os dados das planilhas(abas) específicas
dados = wb['DadosParticipante']
simulacao = wb1['Calculo']
salarios = wb2['Salarios']
 
    # Criando variáveis que recebem os índices das colunas para a iteração no loop
COLUNA_CPF = 0
COLUNA_NOME = 1
COLUNA_DT_NASC = 2
COLUNA_DT_ADESAO = 3
COLUNA_DT_DESLIG = 4
COLUNA_TERMCONTRIB = 5
COLUNA_TEMPSERVANTERIOR = 6
    
    # Itera sobre as linhas, ou seja, para cada linha em dados acontece a iteração começando pela linha 2 (pulando os cabeçalhos)
for row in dados.iter_rows(min_row=2):
    #print(row[0].value)
    # Em cada linha abaixo o \"objeto\" que é a célula selecionada recebe o valor que consta na célula que está sendo iterada no momento baseada no índice da coluna
    simulacao['O3'].value = row[COLUNA_CPF].value
    simulacao['C3'].value = row[COLUNA_NOME].value
    simulacao['C4'].value = row[COLUNA_DT_NASC].value
    simulacao['C5'].value = row[COLUNA_DT_ADESAO].value
    simulacao['C6'].value = row[COLUNA_DT_DESLIG].value
    simulacao['C7'].value = row[COLUNA_TERMCONTRIB].value
    simulacao['O6'].value = row[COLUNA_TEMPSERVANTERIOR].value
    
    cont = 17

    for linha in salarios.iter_rows(min_row=1, max_row=36):
        salario = linha[0].value
        simulacao['I%d' % cont].value = salario
        cont += 1
    #wb1.save('C:/Users/pedro/Desktop/Automatizacao/%s.xlsx' % row[COLUNA_NOME].value)
            
    wb1.save('C:/Users/pedro/Desktop/Automatizacao/%s.xlsx' % row[COLUNA_NOME].value)